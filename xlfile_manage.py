# install package
    # pip install openpyxl
# put the xl file to app that need to assess the data
import openpyxl as xl

wb = xl.load_workbook('xlfile_name.xlsl')
sheet = wb['Sheet1'] #specify the sheet
#cell = sheet['a1'] #a1 cell
#cell = sheet.cell(1 ,1) # a1 cell
#sheet.max_row # to get the maximum length of row

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    # print(cell.value) # to print value
    corrented_price = cell.value * .7
    corrented_price_cell = sheet.cell(row, 4)
    corrented_price_cell.value = corrented_price

wb.save('name.xlsx')




# Excel Export using xlwt
# pip install xlwt


def display_value(choices, field):
    options = [
        When(**{field: k, 'then': Value(v)})
        for k, v in choices
    ]
    return Case(
        *options, output_field=CharField()
    )


def make_heading(name):
    return name.capitalize().replace('_display', '').replace('__date', '').replace('__', ' ').replace('_', ' ')



def employer_export_excel(request):
    instances = Employer.objects.filter(is_deleted=False)
    my_field = {
        'payroll__month_display': display_value(MONTH_CHOICES, 'payroll__month'),
    }
    list_data = EmployerSerializerData(instance=instances,many=True,context={'request':request}).data
    wb = xlwt.Workbook()
    title = "Employer report"
    ws = wb.add_sheet(title)

    for index, item in enumerate(list_data):
        if index == 0:
            for key_index, key in enumerate(item.keys()):
                ws.write(index, key_index, make_heading(key))

        for item_index, value in enumerate(item.values()):
            ws.write(index+1, item_index, str(value))

    if not os.path.isdir(str(SETTINGS.MEDIA_ROOT) + "/excels/"):
        os.makedirs(str(SETTINGS.MEDIA_ROOT) + "/excels/")

    media_root = str(SETTINGS.MEDIA_ROOT) + '/excels/' + title + '.xls'
    wb.save(media_root)

    file_url = str(SETTINGS.MEDIA_URL + '/excels/' + title + '.xls')

    return HttpResponseRedirect(file_url)